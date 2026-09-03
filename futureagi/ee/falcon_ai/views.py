import io
import uuid

from django.db.models import Count, Max
from drf_yasg import openapi
from drf_yasg.utils import no_body, swagger_auto_schema
from rest_framework import status
from rest_framework.parsers import MultiPartParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from ee.falcon_ai.models import Conversation, FalconFile, Message
from ee.falcon_ai.serializers import (
    ConversationDetailSerializer,
    ConversationListSerializer,
    MessageFeedbackSerializer,
)
from ee.falcon_ai.serializers_contracts import (
    ConversationCreateRequestSerializer,
    ConversationDetailResponseSerializer,
    ConversationListResponseSerializer,
    ConversationUpdateRequestSerializer,
    FalconErrorResponseSerializer,
    FileUploadResponseSerializer,
    MessageFeedbackResponseSerializer,
    StreamStatusResponseSerializer,
)
from tfc.utils.api_contracts import validated_request
from tfc.utils.general_methods import GeneralMethods

FALCON_FILE_UPLOAD_PARAMETERS = [
    openapi.Parameter(
        "file",
        openapi.IN_FORM,
        type=openapi.TYPE_FILE,
        required=True,
        description="File to upload into the Falcon conversation context.",
    )
]

_gm = GeneralMethods()


class StreamStatusView(APIView):
    """Check if there is an active or recent agent stream for a conversation."""

    permission_classes = [IsAuthenticated]

    @swagger_auto_schema(
        responses={
            200: StreamStatusResponseSerializer,
            403: FalconErrorResponseSerializer,
            404: FalconErrorResponseSerializer,
        }
    )
    def get(self, request, conversation_id):
        from ee.falcon_ai.stream_buffer import StreamBuffer

        # Verify conversation belongs to this user
        organization = getattr(request, "organization", None)
        workspace = getattr(request, "workspace", None)
        if not organization:
            return _gm.forbidden_response("No organization context")

        filters = {
            "id": conversation_id,
            "user": request.user,
            "organization": organization,
        }
        if workspace:
            filters["workspace"] = workspace

        try:
            Conversation.objects.get(**filters)
        except Conversation.DoesNotExist:
            return _gm.not_found("Conversation not found")

        buffer = StreamBuffer(conversation_id)
        stream_status = buffer.get_status()
        return Response(
            {
                "status": True,
                "result": {
                    "stream_status": stream_status or "none",
                },
            }
        )


class ConversationListView(APIView):
    """List and create conversations."""

    permission_classes = [IsAuthenticated]

    @swagger_auto_schema(
        responses={
            200: ConversationListResponseSerializer,
            403: FalconErrorResponseSerializer,
        }
    )
    def get(self, request):
        organization = getattr(request, "organization", None)
        workspace = getattr(request, "workspace", None)

        if not organization:
            return _gm.forbidden_response("No organization context")

        # `metadata__hidden=True` against an empty JSONB `{}` returns NULL in
        # Postgres, which makes `exclude(...)` drop every row (NULL is not TRUE
        # but also not FALSE). Use `metadata__contains={"hidden": True}` which
        # compiles to `@>` — returns a proper boolean.
        conversations = (
            Conversation.objects.filter(
                user=request.user,
                organization=organization,
            )
            .exclude(metadata__contains={"hidden": True})
            .annotate(
                message_count=Count("messages"),
                last_message_at=Max("messages__created_at"),
            )
            .order_by("-updated_at")
        )

        # Filter by workspace if available
        if workspace:
            conversations = conversations.filter(workspace=workspace)

        # Search by title
        search = request.query_params.get("search", "").strip()
        if search:
            conversations = conversations.filter(title__icontains=search)

        # Pagination
        try:
            limit = min(int(request.query_params.get("limit", 20)), 100)
        except (ValueError, TypeError):
            limit = 20
        try:
            offset = max(int(request.query_params.get("offset", 0)), 0)
        except (ValueError, TypeError):
            offset = 0

        total = conversations.count()
        page = conversations[offset : offset + limit]

        serializer = ConversationListSerializer(page, many=True)
        return Response(
            {
                "status": True,
                "results": serializer.data,
                "total": total,
                "limit": limit,
                "offset": offset,
                "has_more": (offset + limit) < total,
            }
        )

    @validated_request(
        ConversationCreateRequestSerializer,
        responses={
            201: ConversationDetailResponseSerializer,
            403: FalconErrorResponseSerializer,
        },
    )
    def post(self, request):
        organization = getattr(request, "organization", None)
        workspace = getattr(request, "workspace", None)

        if not organization:
            return _gm.forbidden_response("No organization context")

        data = request.validated_data
        title = data.get("title") or "New conversation"
        context_page = data.get("context_page", "")
        # Internal/embedded conversations (Fix-tab RCA) ask to be hidden from
        # the chat-history list via the metadata.hidden mechanism the list view
        # already excludes on.
        metadata = {"hidden": True} if data.get("hidden") else {}

        conversation = Conversation.objects.create(
            user=request.user,
            organization=organization,
            workspace=workspace,
            title=title,
            context_page=context_page,
            metadata=metadata,
        )

        serializer = ConversationDetailSerializer(conversation)
        return Response(
            {"status": True, "result": serializer.data},
            status=status.HTTP_201_CREATED,
        )


class ConversationDetailView(APIView):
    """Get, update, or delete a conversation."""

    permission_classes = [IsAuthenticated]

    def _get_conversation(self, request, conversation_id):
        organization = getattr(request, "organization", None)
        workspace = getattr(request, "workspace", None)
        if not organization:
            return None, _gm.forbidden_response("No organization context")

        filters = {
            "id": conversation_id,
            "user": request.user,
            "organization": organization,
        }
        if workspace:
            filters["workspace"] = workspace

        try:
            conversation = Conversation.objects.get(**filters)
            return conversation, None
        except Conversation.DoesNotExist:
            return None, _gm.not_found("Conversation not found")

    @swagger_auto_schema(
        responses={
            200: ConversationDetailResponseSerializer,
            403: FalconErrorResponseSerializer,
            404: FalconErrorResponseSerializer,
        }
    )
    def get(self, request, conversation_id):
        conversation, error = self._get_conversation(request, conversation_id)
        if error:
            return error

        serializer = ConversationDetailSerializer(conversation)
        return Response({"status": True, "result": serializer.data})

    @validated_request(
        ConversationUpdateRequestSerializer,
        responses={
            200: ConversationDetailResponseSerializer,
            403: FalconErrorResponseSerializer,
            404: FalconErrorResponseSerializer,
        },
    )
    def patch(self, request, conversation_id):
        conversation, error = self._get_conversation(request, conversation_id)
        if error:
            return error

        data = request.validated_data
        if "title" in data:
            conversation.title = data["title"]
            conversation.save(update_fields=["title", "updated_at"])

        serializer = ConversationDetailSerializer(conversation)
        return Response({"status": True, "result": serializer.data})

    @swagger_auto_schema(
        responses={
            204: "Conversation deleted",
            403: FalconErrorResponseSerializer,
            404: FalconErrorResponseSerializer,
        }
    )
    def delete(self, request, conversation_id):
        conversation, error = self._get_conversation(request, conversation_id)
        if error:
            return error

        # Soft delete via BaseModel.delete()
        conversation.delete()
        return Response({"status": True}, status=status.HTTP_204_NO_CONTENT)


class MessageFeedbackView(APIView):
    """Update feedback on a message."""

    permission_classes = [IsAuthenticated]

    @validated_request(
        request_serializer=MessageFeedbackSerializer,
        responses={
            200: MessageFeedbackResponseSerializer,
            403: FalconErrorResponseSerializer,
            404: FalconErrorResponseSerializer,
        },
        reject_unknown_fields=True,
    )
    def post(self, request, message_id):
        organization = getattr(request, "organization", None)
        workspace = getattr(request, "workspace", None)
        if not organization:
            return _gm.forbidden_response("No organization context")

        filters = {
            "id": message_id,
            "conversation__user": request.user,
            "conversation__organization": organization,
        }
        if workspace:
            filters["conversation__workspace"] = workspace

        try:
            message = Message.objects.select_related("conversation").get(**filters)
        except Message.DoesNotExist:
            return _gm.not_found("Message not found")

        message.feedback = request.validated_data["feedback"]
        message.save(update_fields=["feedback", "updated_at"])

        return Response({"status": True, "result": {"feedback": message.feedback}})


class FileUploadView(APIView):
    """Upload a file for use in Falcon AI conversations."""

    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser]

    MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB
    ALLOWED_TYPES = {
        "text/plain",
        "text/csv",
        "text/html",
        "text/markdown",
        "application/json",
        "application/pdf",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "image/png",
        "image/jpeg",
        "image/gif",
        "image/webp",
    }

    # Text-based types where we can extract content for LLM context
    TEXT_TYPES = {
        "text/plain",
        "text/csv",
        "text/html",
        "text/markdown",
        "application/json",
    }

    @swagger_auto_schema(
        request_body=no_body,
        manual_parameters=FALCON_FILE_UPLOAD_PARAMETERS,
        runtime_request_validation=True,
        responses={
            201: FileUploadResponseSerializer,
            400: FalconErrorResponseSerializer,
            403: FalconErrorResponseSerializer,
        },
    )
    def post(self, request):
        file = request.FILES.get("file")
        if not file:
            return _gm.bad_request("No file provided")

        if file.size > self.MAX_FILE_SIZE:
            return _gm.bad_request("File too large (max 10MB)")

        org = getattr(request, "organization", None)
        workspace = getattr(request, "workspace", None)

        if not org:
            return _gm.forbidden_response("No organization context")

        import os
        import re
        from urllib.parse import urlparse

        from tfc.settings.settings import MINIO_URL, UPLOAD_BUCKET_NAME

        # Sanitize filename to prevent path traversal
        if not re.match(r"^[\w\s\-\.()]+$", file.name):
            file_name = re.sub(r"[^\w\s\-\.()]", "_", file.name)
        else:
            file_name = file.name

        object_key = f"falcon-ai/{org.id}/{uuid.uuid4()}/{file_name}"
        file_bytes = file.read()

        # Detect content type from extension if browser didn't set it.
        content_type = file.content_type or "application/octet-stream"
        if content_type == "application/octet-stream":
            ext_map = {
                ".csv": "text/csv",
                ".txt": "text/plain",
                ".json": "application/json",
                ".html": "text/html",
                ".md": "text/markdown",
                ".pdf": "application/pdf",
                ".png": "image/png",
                ".jpg": "image/jpeg",
                ".jpeg": "image/jpeg",
                ".gif": "image/gif",
                ".webp": "image/webp",
                ".xlsx": (
                    "application/vnd.openxmlformats-officedocument"
                    ".spreadsheetml.sheet"
                ),
                ".docx": (
                    "application/vnd.openxmlformats-officedocument"
                    ".wordprocessingml.document"
                ),
            }
            ext = os.path.splitext(file_name)[1].lower()
            content_type = ext_map.get(ext, content_type)

        if content_type not in self.ALLOWED_TYPES:
            return _gm.bad_request("File type not allowed")

        # Check for dangerous file content signatures
        DANGEROUS_SIGNATURES = [
            b"MZ",  # Windows executables
            b"\x7fELF",  # Linux executables
            b"#!/",  # Shell scripts
            b"PK\x03\x04",  # Could be ZIP/JAR with executables
        ]
        safe_binary_types = (
            "application/pdf",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        )
        file_head = file_bytes[:4]
        for sig in DANGEROUS_SIGNATURES:
            if file_head.startswith(sig) and content_type not in safe_binary_types:
                return _gm.bad_request("File type not allowed")

        from minio import Minio

        minio_endpoint = os.getenv("MINIO_ENDPOINT") or os.getenv(
            "S3_ENDPOINT_URL", "minio:9000"
        )
        parsed_endpoint = urlparse(minio_endpoint)
        if parsed_endpoint.scheme:
            minio_endpoint = parsed_endpoint.netloc
            secure = parsed_endpoint.scheme == "https"
        else:
            secure = "amazonaws" in minio_endpoint

        access_key = (
            os.getenv("MINIO_ROOT_USER")
            or os.getenv("S3_ACCESS_KEY")
            or os.getenv("AWS_ACCESS_KEY_ID", "")
        )
        secret_key = (
            os.getenv("MINIO_ROOT_PASSWORD")
            or os.getenv("S3_SECRET_KEY")
            or os.getenv("AWS_SECRET_ACCESS_KEY", "")
        )

        minio_client = Minio(
            minio_endpoint,
            access_key=access_key,
            secret_key=secret_key,
            secure=secure,
        )

        # Ensure bucket exists
        if not minio_client.bucket_exists(UPLOAD_BUCKET_NAME):
            minio_client.make_bucket(UPLOAD_BUCKET_NAME)

        minio_client.put_object(
            UPLOAD_BUCKET_NAME,
            object_key,
            io.BytesIO(file_bytes),
            len(file_bytes),
            content_type=content_type,
        )

        storage_url = f"{MINIO_URL}/{UPLOAD_BUCKET_NAME}/{object_key}"

        # Extract text content for LLM context
        text_content = ""
        is_text = content_type in self.TEXT_TYPES or file_name.endswith(
            (".csv", ".txt", ".json", ".md", ".html")
        )
        if is_text:
            try:
                text_content = file_bytes.decode("utf-8", errors="replace")[:50000]
            except Exception:
                pass

        # Excel (.xlsx) extraction
        elif content_type in (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ) or file_name.endswith(".xlsx"):
            try:
                import openpyxl

                wb = openpyxl.load_workbook(
                    io.BytesIO(file_bytes), read_only=True, data_only=True
                )
                sheets_text = []
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    rows = []
                    for row in ws.iter_rows(
                        max_row=200, values_only=True
                    ):  # Limit to 200 rows
                        rows.append(
                            [str(cell) if cell is not None else "" for cell in row]
                        )
                    if rows:
                        # First row as headers
                        header = ",".join(rows[0])
                        data_rows = [",".join(r) for r in rows[1:]]
                        sheet_text = f"## Sheet: {sheet_name}\n{header}\n" + "\n".join(
                            data_rows
                        )
                        sheets_text.append(sheet_text)
                wb.close()
                text_content = "\n\n".join(sheets_text)[:50000]
            except ImportError:
                text_content = "Excel parsing unavailable (openpyxl not installed)"
            except Exception as e:
                text_content = f"Failed to parse Excel: {str(e)}"

        # PDF extraction
        elif content_type == "application/pdf" or file_name.endswith(".pdf"):
            try:
                import pdfplumber

                with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
                    pages_text = []
                    for i, page in enumerate(pdf.pages[:50]):  # Limit to 50 pages
                        page_text = page.extract_text()
                        if page_text:
                            pages_text.append(f"--- Page {i + 1} ---\n{page_text}")
                    text_content = "\n\n".join(pages_text)[:50000]
            except ImportError:
                text_content = "PDF parsing unavailable (pdfplumber not installed)"
            except Exception as e:
                text_content = f"Failed to parse PDF: {str(e)}"

        # DOCX extraction
        elif (
            content_type
            == "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
            or file_name.endswith(".docx")
        ):
            try:
                import docx

                doc = docx.Document(io.BytesIO(file_bytes))
                paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
                text_content = "\n".join(paragraphs)[:50000]
            except ImportError:
                text_content = "DOCX parsing unavailable (python-docx not installed)"
            except Exception as e:
                text_content = f"Failed to parse DOCX: {str(e)}"

        falcon_file = FalconFile.objects.create(
            organization=org,
            workspace=workspace,
            user=request.user,
            name=file_name,
            size=file.size,
            content_type=content_type,
            storage_key=object_key,
            storage_url=storage_url,
            text_content=text_content,
        )

        return Response(
            {
                "status": True,
                "result": {
                    "id": str(falcon_file.id),
                    "name": falcon_file.name,
                    "size": falcon_file.size,
                    "content_type": falcon_file.content_type,
                    "url": storage_url,
                },
            },
            status=201,
        )
